from flask import Flask, render_template, request, send_file
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import cm
from reportlab.lib.utils import ImageReader
from io import BytesIO
import os
import zipfile
from PIL import Image, ImageDraw, ImageFont
from pypdf import PdfReader, PdfWriter
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = os.environ.get('SESSION_SECRET', 'dev-secret-key')

pdfmetrics.registerFont(TTFont('LibreBaskerville', 'fonts/LibreBaskerville-Regular.ttf'))
pdfmetrics.registerFont(TTFont('PinyonScript', 'fonts/PinyonScript-Regular.ttf'))

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/preview', methods=['POST'])
def preview_certificate():
    data = request.get_json()
    nama = data.get('nama', '')
    kelas = data.get('kelas', '')
    peringkat = data.get('peringkat', '')
    wali_kelas = data.get('wali_kelas', '')
    pimpinan_ponpes = data.get('pimpinan_ponpes', '')
    
    # Get custom positions if provided
    nama_y = float(data.get('nama_y', 310))
    nama_x = float(data.get('nama_x')) if data.get('nama_x') else None
    peringkat_y = float(data.get('peringkat_y', 235))
    peringkat_x = float(data.get('peringkat_x')) if data.get('peringkat_x') else None
    wali_y = float(data.get('wali_y', 100))
    wali_x = float(data.get('wali_x')) if data.get('wali_x') else None
    pimpinan_y = float(data.get('pimpinan_y', 100))
    pimpinan_x = float(data.get('pimpinan_x')) if data.get('pimpinan_x') else None
    
    pdf_buffer = create_single_certificate(nama, kelas, peringkat, wali_kelas, pimpinan_ponpes,
                                          nama_y, nama_x, peringkat_y, peringkat_x, 
                                          wali_y, wali_x, pimpinan_y, pimpinan_x)
    
    return send_file(
        pdf_buffer,
        mimetype='application/pdf',
        as_attachment=False,
        download_name='preview.pdf'
    )

def draw_centered_text(c, text, x, y, font_name="LibreBaskerville",
                       font_size=12, color_rgb=(0, 0, 0)):
    c.setFont(font_name, font_size)
    c.setFillColorRGB(*[v/255 for v in color_rgb])  # pastikan pakai 0–1

    if x is None:
        page_width = 842  # A4 landscape
        x = page_width / 2

    text_width = c.stringWidth(text, font_name, font_size)
    c.drawString(x - text_width / 2, y, text)


def create_text_image(text, font_path, font_size, color=(0, 0, 0)):
    font = ImageFont.truetype(font_path, font_size)
    
    dummy_img = Image.new('RGBA', (1, 1))
    dummy_draw = ImageDraw.Draw(dummy_img)
    bbox = dummy_draw.textbbox((0, 0), text, font=font)
    text_width = int(bbox[2] - bbox[0])
    text_height = int(bbox[3] - bbox[1])
    
    img = Image.new('RGBA', (text_width + 20, text_height + 20), (255, 255, 255, 0))
    draw = ImageDraw.Draw(img)
    draw.text((10, 10), text, font=font, fill=color)
    
    return img

def create_single_certificate(nama, kelas, peringkat, wali_kelas, pimpinan_ponpes, 
                             nama_y=310, nama_x=None, peringkat_y=235, peringkat_x=None, 
                             wali_y=100, wali_x=None, pimpinan_y=100, pimpinan_x=None):
    template_pdf = PdfReader('static/desain/desain-1.pdf')
    template_page = template_pdf.pages[0]
    
    page_width = float(template_page.mediabox.width)
    page_height = float(template_page.mediabox.height)
    
    overlay_buffer = BytesIO()
    c = canvas.Canvas(overlay_buffer, pagesize=(page_width, page_height))
    
    # Warna RGB
    gold_color_pil = (184, 134, 11)
    dark_text_color_pil = (31, 52, 83)
    gold_color_pdf = tuple(v / 255 for v in gold_color_pil)
    dark_text_color_pdf = tuple(v / 255 for v in dark_text_color_pil)

    # === Nama besar ===
    nama_img = create_text_image(nama, "fonts/OPTIEngraversOldEnglish.otf", 85, gold_color_pil)
    nama_img_buffer = BytesIO()
    nama_img.save(nama_img_buffer, format='PNG')
    nama_img_buffer.seek(0)
    
    nama_img_width = nama_img.width * 0.5
    nama_img_height = nama_img.height * 0.5

    # Titik tengah nama
    nama_center_x = nama_x if nama_x is not None else page_width / 2
    nama_x_pos = nama_center_x - nama_img_width / 2

    c.drawImage(
        ImageReader(nama_img_buffer),
        nama_x_pos,
        nama_y,
        width=nama_img_width,
        height=nama_img_height,
        mask='auto'
    )

    # === Peringkat ===
    peringkat_text = f"peringkat ke {peringkat} kelas {kelas}"
    draw_centered_text(
        c,
        peringkat_text,
        peringkat_x or page_width / 2,
        peringkat_y,
        font_name="LibreBaskerville",
        font_size=12,
        color_rgb=gold_color_pdf
    )

    # === Nama wali kelas (centered based on wali_x) ===
    c.setFont("LibreBaskerville", 13)
    c.setFillColorRGB(*dark_text_color_pdf)
    wali_text_width = c.stringWidth(wali_kelas, "LibreBaskerville", 13)
    wali_center_x = wali_x if wali_x is not None else page_width / 2 - 150  # default agak kiri
    c.drawString(wali_center_x - wali_text_width / 2, wali_y, wali_kelas)

    # === Nama pimpinan ponpes (centered based on pimpinan_x) ===
    c.setFont("LibreBaskerville", 13)
    c.setFillColorRGB(*dark_text_color_pdf)
    pimpinan_text_width = c.stringWidth(pimpinan_ponpes, "LibreBaskerville", 13)
    pimpinan_center_x = pimpinan_x if pimpinan_x is not None else page_width / 2 + 150  # default agak kanan
    c.drawString(pimpinan_center_x - pimpinan_text_width / 2, pimpinan_y, pimpinan_ponpes)

    c.save()
    
    overlay_buffer.seek(0)
    overlay_pdf = PdfReader(overlay_buffer)
    overlay_page = overlay_pdf.pages[0]
    
    template_page.merge_page(overlay_page)
    
    output = PdfWriter()
    output.add_page(template_page)
    
    final_buffer = BytesIO()
    output.write(final_buffer)
    final_buffer.seek(0)
    
    return final_buffer

@app.route('/generate', methods=['POST'])
def generate_certificate():
    nama = request.form.get('nama', '')
    kelas = request.form.get('kelas', '')
    peringkat = request.form.get('peringkat', '')
    wali_kelas = request.form.get('wali_kelas', '')
    pimpinan_ponpes = request.form.get('pimpinan_ponpes', '')
    
    pdf_buffer = create_single_certificate(nama, kelas, peringkat, wali_kelas, pimpinan_ponpes)
    
    return send_file(
        pdf_buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'Piagam_Penghargaan_{nama.replace(" ", "_")}.pdf'
    )

def safe_float(value, default=None):
    """Parse float safely (return default if blank or invalid)."""
    try:
        if value is None or str(value).strip() == "":
            return default
        return float(value)
    except ValueError:
        return default


@app.route('/generate-bulk', methods=['POST'])
def generate_bulk():
    if 'excel_file' not in request.files:
        return "No file uploaded", 400

    excel_file = request.files['excel_file']
    if excel_file.filename == '':
        return "No file selected", 400

    # Gunakan safe_float agar parsing stabil & konsisten
    nama_y = safe_float(request.form.get('bulk_nama_y'), 310)
    nama_x = safe_float(request.form.get('bulk_nama_x'))
    peringkat_y = safe_float(request.form.get('bulk_peringkat_y'), 235)
    peringkat_x = safe_float(request.form.get('bulk_peringkat_x'))
    wali_y = safe_float(request.form.get('bulk_wali_y'), 100)
    wali_x = safe_float(request.form.get('bulk_wali_x'))
    pimpinan_y = safe_float(request.form.get('bulk_pimpinan_y'), 100)
    pimpinan_x = safe_float(request.form.get('bulk_pimpinan_x'))

    
    excel_buffer = BytesIO(excel_file.read())
    wb = load_workbook(excel_buffer)
    ws = wb.active
    
    # Create ZIP in memory with better compression
    zip_buffer = BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED, compresslevel=6) as zip_file:
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                break
            
            nama = str(row[0]) if row[0] else ''
            kelas = str(row[1]) if row[1] else ''
            peringkat = str(row[2]) if row[2] else ''
            wali_kelas = str(row[3]) if row[3] else ''
            pimpinan_ponpes = str(row[4]) if row[4] else ''
            
            if nama and kelas and peringkat:
                pdf_buffer = create_single_certificate(
                    nama, kelas, peringkat, wali_kelas, pimpinan_ponpes,
                    nama_y, nama_x, peringkat_y, peringkat_x, 
                    wali_y, wali_x, pimpinan_y, pimpinan_x
                )
                
                safe_nama = nama.replace(" ", "_").replace("/", "_")
                filename = f'Piagam_{safe_nama}_{kelas}.pdf'
                
                zip_file.writestr(filename, pdf_buffer.getvalue())
                row_count += 1
                
                # Clear buffer to free memory
                pdf_buffer.close()
    
    zip_buffer.seek(0)
    
    # Use send_file with proper headers to avoid Chrome blocking
    response = send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name='Piagam_Penghargaan_Bulk.zip'
    )
    
    # Add headers to prevent blocking
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['Content-Disposition'] = 'attachment; filename=Piagam_Penghargaan_Bulk.zip'
    
    return response

if __name__ == '__main__':
    debug_mode = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    app.run(host='0.0.0.0', port=5000, debug=debug_mode)
